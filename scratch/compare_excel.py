import openpyxl
import os
import json

wb1 = openpyxl.load_workbook('manjul product data.xlsx')
sheet1 = wb1.active
rows1 = list(sheet1.iter_rows(values_only=True))

wb2 = openpyxl.load_workbook('manjul sheet 1 - Copy (1).xlsx')
sheet2 = wb2.active
rows2 = list(sheet2.iter_rows(values_only=True))

print("=== Old Excel (manjul product data.xlsx) ===")
print("Headers:", rows1[0])
print(f"Total rows: {len(rows1)}")

print("\n=== New Excel (manjul sheet 1 - Copy (1).xlsx) ===")
print("Headers:", rows2[0])
print(f"Total rows: {len(rows2)}")

print("\nComparing Row by Row:")
for i in range(1, max(len(rows1), len(rows2))):
    r1_name = rows1[i][0] if i < len(rows1) and rows1[i] else 'N/A'
    r2_name = rows2[i][0] if i < len(rows2) and rows2[i] else 'N/A'
    excel_row_num = i + 1
    image_num = excel_row_num # since row 2 is 2.jpeg
    print(f"Excel Row {excel_row_num} ({image_num}.jpeg):")
    print(f"   Old: {r1_name}")
    print(f"   New: {r2_name}")
